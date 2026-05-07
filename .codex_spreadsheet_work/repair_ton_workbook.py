from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


ROOT = Path("/Users/victorsim/Desktop/Projects/learn_more_v1.0")
SRC = ROOT / "outputs" / "ton_rollover_model" / "TON_rolling_position_calculator.xlsx"
TMP = ROOT / "outputs" / "ton_rollover_model" / "TON_rolling_position_calculator_repaired.xlsx"


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


wb = load_workbook(SRC)

# The original file's chart/drawing XML is readable by libraries, but can trigger
# Excel compatibility repair on some desktop Excel versions. Keep the chart data,
# remove embedded drawing objects, and let the workbook stay formula-driven.
for ws in wb.worksheets:
    ws._charts = []
    ws._images = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and len(cell.value) > 28:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

set_widths(
    wb["摘要"],
    {
        "A": 18,
        "B": 13,
        "C": 9,
        "D": 24,
        "E": 3,
        "F": 5,
        "G": 15,
        "H": 42,
        "I": 3,
        "J": 3,
    },
)

set_widths(wb["参数"], {"A": 22, "B": 12, "C": 11, "D": 38})

for name in ["U本位滚仓", "币本位滚仓", "风险预算滚仓"]:
    ws = wb[name]
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 11
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["R"].width = 10
    if name == "币本位滚仓":
        ws.column_dimensions["T"].width = 10
    if name == "风险预算滚仓":
        ws.column_dimensions["S"].width = 12

set_widths(wb["敏感性分析"], {"A": 16, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10, "G": 24, "H": 12})
set_widths(wb["检查"], {"A": 24, "B": 12, "C": 14, "D": 12, "E": 10, "F": 36})

for ws in wb.worksheets:
    for row in range(1, min(ws.max_row, 40) + 1):
        ws.row_dimensions[row].height = 20
    ws.freeze_panes = "A4" if ws.title in {"摘要", "参数", "检查"} else "A5"

wb.save(TMP)
TMP.replace(SRC)
print(SRC)
