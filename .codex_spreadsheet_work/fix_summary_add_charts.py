from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path("/Users/victorsim/Desktop/Projects/learn_more_v1.0")
P = ROOT / "outputs" / "ton_rollover_model" / "TON_rolling_position_calculator.xlsx"

wb = load_workbook(P)

navy = "12355B"
blue = "DBEAFE"
soft = "F8FAFC"
border_color = "CBD5E1"
white = "FFFFFF"
thin = Side(style="thin", color=border_color)
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cell(cell, fill=None, bold=False, color="000000", align="center"):
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Aptos", size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = border


# 1) Make risk-budget formulas more Excel-compatible.
rb = wb["风险预算滚仓"]
for r in range(5, 17):
    if r == 5:
        rb[f"D{r}"] = '=IF($A5="","",\'参数\'!$B$5)'
        rb[f"Q{r}"] = '=IF($A5="", "", 0)'
    else:
        prev = r - 1
        rb[f"D{r}"] = (
            f'=IF($A{r}="","",\'参数\'!$B$5+SUM($H$5:H{prev})*B{r}'
            f'-SUMPRODUCT($H$5:H{prev},$C$5:C{prev})-K{prev})'
        )
        rb[f"Q{r}"] = (
            f'=IF($A{r}="","",SUM($H$5:H{prev})*B{r}'
            f'-SUMPRODUCT($H$5:H{prev},$C$5:C{prev})-K{prev})'
        )
    rb[f"R{r}"] = (
        f'=IF($A{r}="","",\'参数\'!$B$5+SUM($H$5:H{r})*\'参数\'!$B$7'
        f'-SUMPRODUCT($H$5:H{r},$C$5:C{r})-K{r}'
        f'-SUM($H$5:H{r})*\'参数\'!$B$7*\'参数\'!$B$13)'
    )

# 2) Replace error-prone summary LOOKUP with direct INDEX based on max add count.
dash = wb["摘要"]
dash["B14"] = "=INDEX('风险预算滚仓'!R5:R16,MIN('参数'!$B$29,ROWS('风险预算滚仓'!R5:R16)))"
dash["B15"] = "=B14/'参数'!$B$5"
dash["B14"].number_format = "$#,##0"
dash["B15"].number_format = "0.0x"

# 3) Rebuild chart data block.
for row in range(20, 28):
    for col in range(1, 5):
        c = dash.cell(row, col)
        c.value = None
        style_cell(c, fill=white)

dash["A20"] = "图表数据"
dash["B20"] = "最终倍数"
dash["C20"] = "强平/风险次数"
dash["D20"] = "说明"
for c in dash[20][0:4]:
    style_cell(c, fill=navy, bold=True, color=white)

chart_rows = [
    ("现货", "=B5", 0, "不使用杠杆"),
    ("U本位", "=B7", "=B8", "固定目标杠杆复投"),
    ("币本位", "=B11", "=B12", "币本位固定目标杠杆"),
    ("风险预算", "=B15", 0, "分笔加仓，控制单笔风险"),
    ("目标币仓降杠杆", "=B17", 0, "盈利后降杠杆，维持目标币仓"),
]
for i, row in enumerate(chart_rows, 21):
    for j, value in enumerate(row, 1):
        dash.cell(i, j, value)
        style_cell(dash.cell(i, j), fill=soft if i % 2 else white, align="left" if j in {1, 4} else "center")
dash["B21"].number_format = "0.0x"
for r in range(22, 26):
    dash[f"B{r}"].number_format = "0.0x"
    dash[f"C{r}"].number_format = "0"

# 4) Create charts on a dedicated sheet.
if "图表" in wb.sheetnames:
    del wb["图表"]
charts = wb.create_sheet("图表", 1)
charts.sheet_view.showGridLines = False

charts.merge_cells("A1:L1")
charts["A1"] = "滚仓模式图表"
charts["A1"].fill = PatternFill("solid", fgColor=navy)
charts["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=white)
charts["A1"].alignment = Alignment(vertical="center")

charts.merge_cells("A2:L2")
charts["A2"] = "图表引用摘要页和各模式页公式。修改参数后，Excel 打开并重算即可更新。"
charts["A2"].fill = PatternFill("solid", fgColor=blue)
charts["A2"].font = Font(name="Aptos", size=10, color="334155")
charts["A2"].alignment = Alignment(wrap_text=True)

# Chart source tables on 图表 sheet.
charts["A4"] = "模式"
charts["B4"] = "最终倍数"
charts["C4"] = "风险次数"
for c in charts[4][0:3]:
    style_cell(c, fill=navy, bold=True, color=white)
for i, label in enumerate(["现货", "U本位", "币本位", "风险预算", "目标币仓降杠杆"], 5):
    charts.cell(i, 1, label)
    charts.cell(i, 2, f"='摘要'!B{[5,7,11,15,17][i-5]}")
    charts.cell(i, 3, f"='摘要'!B{[8,8,12,8,8][i-5]}" if label in {"U本位", "币本位"} else 0)
    for col in range(1, 4):
        style_cell(charts.cell(i, col), fill=soft if i % 2 else white, align="left" if col == 1 else "center")
    charts.cell(i, 2).number_format = "0.0x"

bar = BarChart()
bar.type = "col"
bar.title = "各模式最终倍数对比"
bar.y_axis.title = "本金倍数"
bar.x_axis.title = "模式"
bar.add_data(Reference(charts, min_col=2, min_row=4, max_row=9), titles_from_data=True)
bar.set_categories(Reference(charts, min_col=1, min_row=5, max_row=9))
bar.height = 8
bar.width = 16
bar.style = 10
charts.add_chart(bar, "E4")

charts["A15"] = "Step"
charts["B15"] = "TON单价"
charts["C15"] = "使用杠杆"
charts["D15"] = "资产总额U"
for c in charts[15][0:4]:
    style_cell(c, fill=navy, bold=True, color=white)
for i in range(12):
    r = 16 + i
    src = 5 + i
    charts.cell(r, 1, f"='目标币仓降杠杆'!A{src}")
    charts.cell(r, 2, f"='目标币仓降杠杆'!C{src}")
    charts.cell(r, 3, f"='目标币仓降杠杆'!G{src}")
    charts.cell(r, 4, f"='目标币仓降杠杆'!J{src}")
    for col in range(1, 5):
        style_cell(charts.cell(r, col), fill=soft if r % 2 else white)
    charts.cell(r, 2).number_format = "$0.0000"
    charts.cell(r, 3).number_format = "0.00x"
    charts.cell(r, 4).number_format = "$#,##0"

line = LineChart()
line.title = "目标币仓降杠杆：价格与杠杆路径"
line.y_axis.title = "价格 / 杠杆"
line.x_axis.title = "Step"
line.add_data(Reference(charts, min_col=2, max_col=3, min_row=15, max_row=27), titles_from_data=True)
line.set_categories(Reference(charts, min_col=1, min_row=16, max_row=27))
line.height = 8
line.width = 16
line.style = 13
charts.add_chart(line, "E19")

for col, width in {"A": 18, "B": 12, "C": 12, "D": 14, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12, "J": 12, "K": 12, "L": 12}.items():
    charts.column_dimensions[col].width = width

# Keep chart drawing count modest and force recalculation on open.
try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
except Exception:
    pass

wb.save(P)
print(P)
