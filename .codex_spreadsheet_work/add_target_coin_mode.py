from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/victorsim/Desktop/Projects/learn_more_v1.0")
P = ROOT / "outputs" / "ton_rollover_model" / "TON_rolling_position_calculator.xlsx"

wb = load_workbook(P)

if "目标币仓降杠杆" in wb.sheetnames:
    del wb["目标币仓降杠杆"]

idx = wb.sheetnames.index("风险预算滚仓") + 1
ws = wb.create_sheet("目标币仓降杠杆", idx)

navy = "12355B"
blue = "DBEAFE"
yellow = "FFF2CC"
soft = "F8FAFC"
border_color = "CBD5E1"
green = "DCFCE7"
red = "FEE2E2"
warn = "FEF3C7"
white = "FFFFFF"

thin = Side(style="thin", color=border_color)
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_range(cells, fill=None, font=None, align=None):
    for row in cells:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            cell.border = border


ws.merge_cells("A1:N1")
ws["A1"] = "目标币仓降杠杆模式"
ws["A1"].fill = PatternFill("solid", fgColor=navy)
ws["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=white)
ws["A1"].alignment = Alignment(vertical="center")

ws.merge_cells("A2:N2")
ws["A2"] = "对应截图里的思路：价格上涨后，利润折算成币本位本金；本金变大后，逐步降低杠杆，目标是维持一个较大的币数量仓位，而不是一直保持固定高杠杆。黄色列可手动改。"
ws["A2"].fill = PatternFill("solid", fgColor=blue)
ws["A2"].font = Font(name="Aptos", size=10, color="334155")
ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")

headers = [
    "Step",
    "本段涨幅",
    "TON单价",
    "币本位本金",
    "手动杠杆",
    "自动目标杠杆",
    "使用杠杆",
    "维持仓位(TON)",
    "本段币本位收益",
    "资产总额U",
    "估算强平价",
    "到强平回撤",
    "安全杠杆上限",
    "状态",
]
for col, h in enumerate(headers, 1):
    c = ws.cell(4, col, h)
    c.fill = PatternFill("solid", fgColor=navy)
    c.font = Font(name="Aptos", size=10, bold=True, color=white)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

default_moves = [0, 0.04, 0.05, 0.05, 0.07, 0.10, 0.20, 0.35, 0.50, 1.00, 2.00, 5.00]
default_lev = [10, 8, 6, 5, 4, 3, 2, 1.5, 1, 0.5, 0.25, 0.1]

for i in range(12):
    r = 5 + i
    prev = r - 1
    ws.cell(r, 1, i)
    ws.cell(r, 2, default_moves[i])
    ws.cell(r, 5, default_lev[i])
    if i == 0:
        ws.cell(r, 3, "='参数'!$B$6")
        ws.cell(r, 4, "='参数'!$B$5/'参数'!$B$6")
        ws.cell(r, 9, "=0")
    else:
        ws.cell(r, 3, f"=C{prev}*(1+B{r})")
        ws.cell(r, 9, f"=H{prev}*(C{r}-C{prev})/C{r}-H{prev}*'参数'!$B$13-H{prev}*'参数'!$B$15*'参数'!$B$16*((C{prev}+C{r})/2)/C{r}")
        ws.cell(r, 4, f"=MAX(0,D{prev}+I{r})")

    # Default target coin position: initial coin principal * 10. The user can override
    # the target by replacing this formula in F2/N2 note or by editing formulas in F rows.
    ws.cell(r, 6, f"=IF(D{r}=0,\"\",('参数'!$B$5/'参数'!$B$6*'参数'!$B$26)/D{r})")
    ws.cell(r, 7, f"=IF(E{r}=\"\",F{r},E{r})")
    ws.cell(r, 8, f"=D{r}*G{r}")
    ws.cell(r, 10, f"=D{r}*C{r}")
    ws.cell(r, 11, f"=IF(G{r}=0,\"\",C{r}*(1-1/G{r})/(1-'参数'!$B$11-'参数'!$B$12))")
    ws.cell(r, 12, f"=IF(K{r}=\"\",\"\",1-K{r}/C{r})")
    ws.cell(r, 13, "='参数'!$B$21")
    ws.cell(r, 14, f"=IF(G{r}>M{r},\"超过安全杠杆\",IF(L{r}<'参数'!$B$17,\"插针风险\",\"OK\"))")

for row in ws.iter_rows(min_row=5, max_row=16, min_col=1, max_col=14):
    for cell in row:
        cell.font = Font(name="Aptos", size=10, color="000000")
        cell.alignment = Alignment(vertical="center")
        cell.border = border
        if cell.column in {2, 5}:
            cell.fill = PatternFill("solid", fgColor=yellow)
            cell.font = Font(name="Aptos", size=10, color="0000FF")
        else:
            cell.fill = PatternFill("solid", fgColor=white)

for r in range(5, 17):
    ws.cell(r, 2).number_format = "0.0%"
    ws.cell(r, 3).number_format = "$0.0000"
    ws.cell(r, 4).number_format = "#,##0.0000"
    ws.cell(r, 5).number_format = "0.00x"
    ws.cell(r, 6).number_format = "0.00x"
    ws.cell(r, 7).number_format = "0.00x"
    ws.cell(r, 8).number_format = "#,##0.0000"
    ws.cell(r, 9).number_format = "#,##0.0000"
    ws.cell(r, 10).number_format = "$#,##0;[Red]($#,##0);-"
    ws.cell(r, 11).number_format = "$0.0000"
    ws.cell(r, 12).number_format = "0.0%"
    ws.cell(r, 13).number_format = "0.00x"

ws.merge_cells("A19:N19")
ws["A19"] = "使用说明"
ws["A19"].fill = PatternFill("solid", fgColor=warn)
ws["A19"].font = Font(name="Aptos", size=11, bold=True, color=navy)
ws["A19"].border = border

notes = [
    "黄色列“本段涨幅”和“手动杠杆”可以直接改，模拟截图里那种 4%、5%、7%、10%、20% 上涨后逐步降杠杆的路径。",
    "币本位本金 = 上一段币本位本金 + 上一段维持仓位带来的币本位收益；费用和资金费会从币本位收益里扣掉。",
    "自动目标杠杆默认按“初始币本金 * 参数页风险预算单笔杠杆”作为目标币仓；如果你想设固定目标仓位，可以直接改 F 列公式。",
    "这个模式和固定目标杠杆滚仓的区别是：它越涨越降杠杆，目标是维持大币数仓位并降低后期爆仓风险。",
]
for i, text in enumerate(notes, 20):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=14)
    c = ws.cell(i, 1, text)
    c.fill = PatternFill("solid", fgColor=soft)
    c.font = Font(name="Aptos", size=10, color="334155")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = border

widths = {
    "A": 7,
    "B": 10,
    "C": 11,
    "D": 13,
    "E": 10,
    "F": 13,
    "G": 10,
    "H": 14,
    "I": 14,
    "J": 13,
    "K": 11,
    "L": 11,
    "M": 12,
    "N": 14,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
for r in range(1, 24):
    ws.row_dimensions[r].height = 21
ws.freeze_panes = "A5"

# Extend summary with this mode without disturbing existing sheets.
dash = wb["摘要"]
dash["A16"] = "目标币仓降杠杆最终资产"
dash["B16"] = "='目标币仓降杠杆'!J16"
dash["C16"] = "USDT"
dash["D16"] = "对应新增模式最后一行资产总额"
dash["A17"] = "目标币仓降杠杆最终倍数"
dash["B17"] = "=B16/'参数'!$B$5"
dash["C17"] = "x"
dash["D17"] = "最终资产 / 初始本金"
for row in range(16, 18):
    for col in range(1, 5):
        cell = dash.cell(row, col)
        cell.border = border
        cell.font = Font(name="Aptos", size=10, color="000000")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if col == 2:
            cell.fill = PatternFill("solid", fgColor=white)
dash["B16"].number_format = "$#,##0"
dash["B17"].number_format = "0.0x"

if "检查" in wb.sheetnames:
    chk = wb["检查"]
    r = chk.max_row + 1
    chk.cell(r, 1, "目标币仓模式杠杆不超过安全上限")
    chk.cell(r, 2, "=MAX('目标币仓降杠杆'!G5:G16)")
    chk.cell(r, 3, "='参数'!$B$21")
    chk.cell(r, 4, f"=B{r}-C{r}")
    chk.cell(r, 5, f'=IF(B{r}<=C{r},"OK","WARN")')
    chk.cell(r, 6, "若 WARN，说明截图式手动杠杆路径在当前最大回撤假设下偏激进。")
    for col in range(1, 7):
        cell = chk.cell(r, col)
        cell.border = border
        cell.font = Font(name="Aptos", size=10, color="000000")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    chk.cell(r, 2).number_format = "0.00x"
    chk.cell(r, 3).number_format = "0.00x"
    chk.cell(r, 4).number_format = "0.00x"

# Keep workbook compatible with Excel by forcing recalculation on open.
try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
except Exception:
    pass

wb.save(P)
print(P)
