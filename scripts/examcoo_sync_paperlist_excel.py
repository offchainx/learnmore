#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
将 examcoo「公共题库中心 -> 学历类 -> 初中教育考试(55303)」的试卷列表，
同步写入本地 Excel（按现有表格格式：科目/试卷编号/试卷名称/总分/题数/时限）。

用法示例：
  python3 scripts/examcoo_sync_paperlist_excel.py \
    --xlsx "/Users/victorsim/Desktop/Examcoo 题目表格.xlsx" \
    --grades "初二,初三,初中会考,中考,其他"

  # 只同步初二（便于先试跑/验收）
  python3 scripts/examcoo_sync_paperlist_excel.py --grades "初二"
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import time
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


BASE_URL = "https://www.examcoo.com"
MID_URL = f"{BASE_URL}/index/detail/mid/1"


SUBJECT_NORMALIZE = {
    "语文": "中文",
    "英语": "英文",
    "生物": "科学",
}


def normalize_grade(raw: str) -> str:
    raw = (raw or "").strip()
    if raw.startswith("初中会考"):
        return "初中会考"
    return raw


def normalize_subject(raw: str) -> str:
    raw = (raw or "").strip()
    return SUBJECT_NORMALIZE.get(raw, raw)


@dataclass(frozen=True)
class SubjectLink:
    grade: str
    subject: str
    k: int

    @property
    def page_url_tpl(self) -> str:
        return f"{BASE_URL}/paperlist/index/k/{self.k}/p/{{page}}"


@dataclass(frozen=True)
class PaperRow:
    subject: str
    paper_id: int
    name: str
    total_score: str
    question_count: str
    duration: str


def http_get(url: str, *, timeout: int = 30) -> str:
    # examcoo 依赖 cookie/session，保持一个 session 更稳。
    raise RuntimeError("use Session.http_get")


class Session:
    def __init__(self, *, delay_s: float) -> None:
        self._s = requests.Session()
        self._delay_s = delay_s
        self._last_req_at = 0.0

        self._s.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/123.0.0.0 Safari/537.36"
                )
            }
        )

    def get(self, url: str, *, timeout: int = 30) -> str:
        now = time.time()
        sleep_for = self._delay_s - (now - self._last_req_at)
        if sleep_for > 0:
            time.sleep(sleep_for)
        r = self._s.get(url, timeout=timeout)
        self._last_req_at = time.time()
        r.raise_for_status()
        r.encoding = "utf-8"
        return r.text


def parse_junior_section_links(html: str) -> List[SubjectLink]:
    soup = BeautifulSoup(html, "html.parser")

    # 找到「初中教育(55303)」这块
    anchor = None
    for a in soup.select("a.catSubAnchor"):
        txt = a.get_text(strip=True)
        if "初中教育" in txt and "55303" in txt:
            anchor = a
            break
    if not anchor:
        raise RuntimeError("未在 /index/detail/mid/1 页面中找到「初中教育(55303)」区块")

    # 该 anchor 所在 catSubBox 之后紧跟着一张 courseTable
    cat_box = anchor.find_parent("div", class_="catSubBox")
    if not cat_box:
        raise RuntimeError("无法定位初中教育(55303)的 catSubBox 容器")

    course_table = cat_box.find("table", class_="courseTable")
    if not course_table:
        raise RuntimeError("无法定位初中教育(55303)的 courseTable")

    links: List[SubjectLink] = []
    for tr in course_table.select("tr"):
        tds = tr.find_all("td", recursive=False)
        if len(tds) < 2:
            continue
        grade_raw = tds[0].get_text(strip=True)
        grade = normalize_grade(grade_raw)
        for a in tds[1].select("a[href]"):
            subject_raw = a.get_text(strip=True)
            subject = normalize_subject(subject_raw)
            href = a.get("href") or ""
            m = re.search(r"/paperlist/index/k/(\d+)/p/(\d+)", href)
            if not m:
                continue
            k = int(m.group(1))
            links.append(SubjectLink(grade=grade, subject=subject, k=k))
    return links


def parse_total_pages(html: str) -> int:
    # 例：总共：6页 154条记录
    m = re.search(r"总共：\s*(\d+)\s*页", html)
    if m:
        return int(m.group(1))
    return 1


def parse_paper_rows(html: str, *, subject: str) -> List[PaperRow]:
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", class_="listGrid")
    if not table:
        return []
    tbody = table.find("tbody")
    if not tbody:
        return []

    rows: List[PaperRow] = []
    for tr in tbody.find_all("tr", recursive=False):
        tds = tr.find_all("td", recursive=False)
        # 列：编号 | icon | 名称 | 总分 | 题数 | 时限 | ...
        if len(tds) < 6:
            continue
        pid_raw = tds[0].get_text(strip=True)
        if not pid_raw.isdigit():
            continue
        paper_id = int(pid_raw)
        name_td = tds[2]
        name = (name_td.get("title") or name_td.get_text(strip=True) or "").strip()
        total_score = (tds[3].get_text(strip=True) or "").strip()
        question_count = (tds[4].get_text(strip=True) or "").strip()
        duration = (tds[5].get_text(strip=True) or "").strip()
        rows.append(
            PaperRow(
                subject=subject,
                paper_id=paper_id,
                name=name,
                total_score=total_score,
                question_count=question_count,
                duration=duration,
            )
        )
    return rows


def ensure_sheet(wb: openpyxl.Workbook, name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(title=name)
    # 保证表头
    headers = ["科目", "试卷编号", "试卷名称", "总分", "题数", "时限"]
    if ws.max_row < 1 or all((ws.cell(1, i).value is None for i in range(1, 7))):
        for i, h in enumerate(headers, start=1):
            ws.cell(1, i).value = h
    else:
        # 若只有一个单元格（空表被保存后常见），也强制修正表头
        first_row = [ws.cell(1, i).value for i in range(1, 7)]
        if first_row != headers:
            for i, h in enumerate(headers, start=1):
                ws.cell(1, i).value = h
    return ws


def load_existing_keys(ws) -> set[Tuple[str, int]]:
    keys: set[Tuple[str, int]] = set()
    for r in range(2, ws.max_row + 1):
        subj = ws.cell(r, 1).value
        pid = ws.cell(r, 2).value
        if not subj or pid is None:
            continue
        try:
            pid_int = int(pid)
        except Exception:
            continue
        keys.add((str(subj), pid_int))
    return keys


def append_rows(ws, rows: Iterable[PaperRow], existing_keys: set[Tuple[str, int]]) -> int:
    appended = 0
    for row in rows:
        key = (row.subject, row.paper_id)
        if key in existing_keys:
            continue
        ws.append(
            [
                ILLEGAL_CHARACTERS_RE.sub("", row.subject or ""),
                row.paper_id,
                ILLEGAL_CHARACTERS_RE.sub("", row.name or ""),
                ILLEGAL_CHARACTERS_RE.sub("", row.total_score or ""),
                ILLEGAL_CHARACTERS_RE.sub("", row.question_count or ""),
                ILLEGAL_CHARACTERS_RE.sub("", row.duration or ""),
            ]
        )
        existing_keys.add(key)
        appended += 1
    return appended


def sync_grade(
    *,
    sess: Session,
    wb: openpyxl.Workbook,
    links_by_grade: Dict[str, List[SubjectLink]],
    grade: str,
    only_subjects: Optional[set[str]],
    max_pages_per_subject: Optional[int],
) -> None:
    ws = ensure_sheet(wb, grade)
    existing_keys = load_existing_keys(ws)
    links = links_by_grade.get(grade, [])
    if only_subjects is not None:
        links = [l for l in links if l.subject in only_subjects]

    print(f"[{grade}] subjects={len(links)} existing_rows={ws.max_row-1}")
    for link in links:
        page1_url = link.page_url_tpl.format(page=1)
        html1 = sess.get(page1_url)
        total_pages = parse_total_pages(html1)
        if max_pages_per_subject is not None:
            total_pages = min(total_pages, max_pages_per_subject)

        added = append_rows(ws, parse_paper_rows(html1, subject=link.subject), existing_keys)
        for p in range(2, total_pages + 1):
            htmlp = sess.get(link.page_url_tpl.format(page=p))
            added += append_rows(ws, parse_paper_rows(htmlp, subject=link.subject), existing_keys)
        print(f"  - {link.subject}: pages={total_pages} +{added}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        default="/Users/victorsim/Desktop/Examcoo 题目表格.xlsx",
        help="Excel 路径",
    )
    ap.add_argument(
        "--grades",
        default="初二,初三,初中会考,中考,其他",
        help="要同步的 sheet 名称（逗号分隔）",
    )
    ap.add_argument(
        "--delay-s",
        type=float,
        default=0.2,
        help="每次 HTTP 请求之间的延迟（秒），避免给对方站点造成压力",
    )
    ap.add_argument(
        "--max-pages-per-subject",
        type=int,
        default=0,
        help="每个科目最多抓取多少页（0 表示不限制，全部抓取）",
    )
    ap.add_argument(
        "--only-subjects",
        default="",
        help="仅同步这些科目（逗号分隔，比如：中文,数学,英文,科学,历史,地理）。留空表示全部。",
    )
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        raise SystemExit(f"Excel 不存在: {args.xlsx}")

    # 轻量备份，避免误覆盖（只备份一次）
    bak = args.xlsx + ".bak"
    if not os.path.exists(bak):
        shutil.copy2(args.xlsx, bak)
        print("backup_saved", bak)

    grades = [g.strip() for g in args.grades.split(",") if g.strip()]
    only_subjects = {s.strip() for s in args.only_subjects.split(",") if s.strip()} or None
    max_pages_per_subject = args.max_pages_per_subject or None

    sess = Session(delay_s=args.delay_s)
    mid_html = sess.get(MID_URL)
    links = parse_junior_section_links(mid_html)
    links_by_grade: Dict[str, List[SubjectLink]] = {}
    for l in links:
        links_by_grade.setdefault(l.grade, []).append(l)

    wb = openpyxl.load_workbook(args.xlsx)

    for g in grades:
        sync_grade(
            sess=sess,
            wb=wb,
            links_by_grade=links_by_grade,
            grade=g,
            only_subjects=only_subjects,
            max_pages_per_subject=max_pages_per_subject,
        )

    wb.save(args.xlsx)
    print("saved", args.xlsx)


if __name__ == "__main__":
    main()
